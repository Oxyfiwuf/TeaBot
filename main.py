# main.py
import asyncio
import re
from datetime import datetime, timedelta
import io
import os
import tempfile
import asyncpg
import zoneinfo

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.types import FSInputFile

from openpyxl.styles import NamedStyle, Font
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from config import BOT_TOKEN, DATABASE_URL, TIMEZONE_STR, DEBUG, MAX_CUSTOM_DAYS, ADMIN_USERS


# Часовий пояс
try:
    TIMEZONE = zoneinfo.ZoneInfo(TIMEZONE_STR)
except zoneinfo.ZoneInfoNotFoundError:
    raise ValueError(f"Невідомий часовий пояс: {TIMEZONE_STR}")

# Глобальний пул з'єднань
pool = None


async def get_pool():
    global pool
    pool = await asyncpg.create_pool(dsn=DATABASE_URL)
    if DEBUG:
        print("Підключено до бази даних PostgreSQL")


async def close_pool():
    global pool
    if pool is not None:
        await pool.close()
        if DEBUG:
            print("Пул з'єднань закрито")


# Функції роботи з БД
async def ensure_user_table(user_id: int):
    table_name = f"tips_{user_id}"
    async with pool.acquire() as conn:
        await conn.execute(f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                date DATE PRIMARY KEY,
                tips INTEGER NOT NULL
            )
        """)


async def upsert_tips(user_id: int, date: datetime, tips: int):
    table_name = f"tips_{user_id}"
    async with pool.acquire() as conn:
        old = await conn.fetchrow(f"SELECT tips FROM {table_name} WHERE date = $1", date.date())
        await conn.execute(
            f"INSERT INTO {table_name} (date, tips) VALUES ($1, $2) "
            f"ON CONFLICT (date) DO UPDATE SET tips = EXCLUDED.tips",
            date.date(), tips
        )
        return old


async def get_tips_in_range(user_id: int, start_date: datetime, end_date: datetime):
    table_name = f"tips_{user_id}"
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT date, tips FROM {table_name} "
            f"WHERE date >= $1 AND date < $2 "
            f"ORDER BY date",
            start_date.date(), end_date.date()
        )
        return [(row['date'], row['tips']) for row in rows]

async def ensure_users_table():
    """Створює таблицю users, якщо її ще немає"""
    async with pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id BIGINT PRIMARY KEY,
                username TEXT,
                first_name TEXT NOT NULL,
                last_name TEXT,
                joined_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
        """)

async def register_user(user):
    """Додає або оновлює запис про користувача в таблиці users"""
    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO users (user_id, username, first_name, last_name)
            VALUES ($1, $2, $3, $4)
            ON CONFLICT (user_id) DO UPDATE
            SET username = EXCLUDED.username,
                first_name = EXCLUDED.first_name,
                last_name = EXCLUDED.last_name
        """,
        user.id,
        user.username,
        user.first_name,
        user.last_name
        )


# FSM стани
class EditSum(StatesGroup):
    waiting_for_date = State()
    waiting_for_sum = State()


class CustomStats(StatesGroup):
    waiting_range = State()
    waiting_days = State()


# Клавіатури
def main_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Статистика", callback_data="stats")],
        [InlineKeyboardButton(text="Додати / змінити суму", callback_data="edit")],
    ])


def stats_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="За цей тиждень", callback_data="current_week"),
         InlineKeyboardButton(text="За останні 7 днів", callback_data="last_7")],
        [InlineKeyboardButton(text="За цей місяць", callback_data="current_month"),
         InlineKeyboardButton(text="За останні 30 днів", callback_data="last_30")],
        [InlineKeyboardButton(text="За цей квартал", callback_data="current_quarter"),
         InlineKeyboardButton(text="За останні 90 днів", callback_data="last_90")],
        [InlineKeyboardButton(text="Від – до (дати)", callback_data="custom_range"),
        InlineKeyboardButton(text="За останні … днів", callback_data="last_n_days")],
        [InlineKeyboardButton(text="За весь час", callback_data="all_time")],
        [InlineKeyboardButton(text="Назад", callback_data="back")],
    ])

# ================== НОВА КЛАВІАТУРА ДЛЯ ВИБОРУ ДАТИ ==================
def date_choice_keyboard():
    from datetime import datetime, timedelta
    import zoneinfo

    TIMEZONE = zoneinfo.ZoneInfo(TIMEZONE_STR)  # використовуємо з config
    today = datetime.now(TIMEZONE).date()
    yesterday = today - timedelta(days=1)

    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text=f"📅 Сьогодні ({today.strftime('%d.%m.%Y')})", callback_data="date_today"),
            InlineKeyboardButton(text=f"📅 Вчора ({yesterday.strftime('%d.%m.%Y')})", callback_data="date_yesterday"),
        ],
        [
            InlineKeyboardButton(text="✍️ Ввести дату вручну", callback_data="date_manual"),
        ],
        [
            InlineKeyboardButton(text="⬅️ Назад", callback_data="back"),
        ]
    ])


# Бот і диспетчер
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()


async def safe_edit_message(cb: CallbackQuery, text: str, reply_markup=None):
    try:
        await cb.message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e).lower():
            await cb.answer()
        else:
            raise


async def calculate_period(user_id: int, start: datetime, end: datetime = None):
    if end is None:
        end = datetime.now(TIMEZONE)
    rows = await get_tips_in_range(user_id, start, end)
    total = sum(tips for _, tips in rows)
    count = len(rows)
    avg = total / count if count else 0
    return total, count, avg


# ================== STARTUP / SHUTDOWN ==================
async def on_startup():
    await get_pool()
    await ensure_users_table()  # гарантуємо, що таблиця users існує при старті
    print("🚀 Бот запущений, база даних підключена")


async def on_shutdown():
    await close_pool()
    print("Бот зупинений")


# Реєструємо події старту і зупинки
dp.startup.register(on_startup)
dp.shutdown.register(on_shutdown)


# ================== HANDLERS ==================
@dp.message(F.text == "/start")
async def start(msg: Message):
    # Створюємо таблицю чайових для користувача
    await ensure_user_table(msg.from_user.id)

    # Реєструємо користувача в загальній таблиці users
    await ensure_users_table()  # створюємо таблицю, якщо немає
    await register_user(msg.from_user)  # додаємо/оновлюємо запис

    await msg.answer("В розділі <i><b>\"Статистика\"</b></i> ви можете переглянути ваші чайові за певний період\nВ розділі <i><b>\"Додати / змінити суму\"</b></i> ви можете записати чайові за певний день (Сьогодні, Вчора або інший день)", reply_markup=main_keyboard(), parse_mode="HTML")


@dp.callback_query(F.data == "back")
async def back(cb: CallbackQuery):
    await safe_edit_message(cb, "Меню", reply_markup=main_keyboard())


@dp.callback_query(F.data == "stats")
async def stats_menu(cb: CallbackQuery):
    await safe_edit_message(cb, "Обери тип статистики:", reply_markup=stats_keyboard())


# === Фіксовані періоди ===
@dp.callback_query(F.data == "current_week")
async def current_week(cb: CallbackQuery):
    now = datetime.now(TIMEZONE)
    start = now - timedelta(days=now.weekday())
    start = start.replace(hour=0, minute=0, second=0, microsecond=0)
    total, count, avg = await calculate_period(cb.from_user.id, start)
    await safe_edit_message(cb,
        f"За цей тиждень (з {start.strftime('%d.%m')})\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн",
        reply_markup=stats_keyboard()
    )


@dp.callback_query(F.data == "current_month")
async def current_month(cb: CallbackQuery):
    now = datetime.now(TIMEZONE)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total, count, avg = await calculate_period(cb.from_user.id, start)
    await safe_edit_message(cb,
        f"За цей місяць ({now.strftime('%B %Y')})\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн",
        reply_markup=stats_keyboard()
    )


@dp.callback_query(F.data == "current_quarter")
async def current_quarter(cb: CallbackQuery):
    now = datetime.now(TIMEZONE)
    quarter_start_month = ((now.month - 1) // 3) * 3 + 1
    start = now.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
    total, count, avg = await calculate_period(cb.from_user.id, start)
    quarter_num = (now.month - 1) // 3 + 1
    await safe_edit_message(cb,
        f"За цей квартал (Q{quarter_num} {now.year})\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн",
        reply_markup=stats_keyboard()
    )


@dp.callback_query(F.data == "all_time")
async def all_time_stats(cb: CallbackQuery):
    user_id = cb.from_user.id
    async with pool.acquire() as conn:
        table_name = f"tips_{user_id}"
        rows = await conn.fetch(f"SELECT date, tips FROM {table_name} ORDER BY date")
    if not rows:
        await safe_edit_message(cb, "За весь час\n\nДаних ще немає", reply_markup=stats_keyboard())
        return
    first_date = rows[0]['date']
    total = sum(r['tips'] for r in rows)
    count = len(rows)
    avg = total / count if count else 0
    await safe_edit_message(cb,
        f"За весь час (з {first_date.strftime('%d.%m.%Y')})\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн",
        reply_markup=stats_keyboard()
    )


@dp.message(F.text == "/a")
async def export_all_data(msg: Message):
    if msg.from_user.id not in ADMIN_USERS:
        return

    await msg.answer("📊 Експортую всі дані... Це може зайняти кілька секунд.")

    try:
        wb = Workbook()
        wb.remove(wb.active)  # видаляємо порожній аркуш

        # Створюємо іменований стиль для дати (формат ДД.ММ.РРРР)
        date_style = NamedStyle(name="date_style", number_format="DD.MM.YYYY")
        wb.add_named_style(date_style)

        async with pool.acquire() as conn:
            tables = await conn.fetch(
                "SELECT tablename FROM pg_tables WHERE schemaname = 'public' AND tablename LIKE 'tips_%'"
            )

            if not tables:
                await msg.answer("⚠️ Даних ще немає.")
                return

            for record in tables:
                table_name = record['tablename']
                user_id = table_name.replace("tips_", "")
                ws = wb.create_sheet(title=f"user_{user_id}")

                # Заголовки
                ws.append(["Дата", "Чайові (грн)"])
                # Форматуємо заголовок жирним (опціонально)
                bold_font = Font(bold=True)

                ws["A1"].font = bold_font
                ws["B1"].font = bold_font

                # Дані
                rows = await conn.fetch(f"SELECT date, tips FROM {table_name} ORDER BY date")
                for row in rows:
                    excel_date = row['date']  # залишаємо як datetime.date
                    ws.append([excel_date, row['tips']])

                # Застосовуємо формат дати до всього стовпця A (крім заголовка)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.style = date_style

                # Автоширина стовпців (для зручності)
                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 15

        # Тимчасовий файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        filename_on_disk = temp_file.name
        temp_file.close()

        wb.save(filename_on_disk)

        input_file = FSInputFile(
            path=filename_on_disk,
            filename=f"tips_export_{datetime.now(TIMEZONE).strftime('%Y-%m-%d')}.xlsx"
        )

        await msg.answer_document(
            document=input_file,
            caption=f"📅 Експорт всіх даних на {datetime.now(TIMEZONE).strftime('%d.%m.%Y')}"
        )

        # Очищаємо тимчасовий файл
        os.remove(filename_on_disk)

    except Exception as e:
        print(f"Помилка експорту: {e}")
        await msg.answer("❌ Помилка при експорті даних.")
        try:
            if 'filename_on_disk' in locals():
                os.remove(filename_on_disk)
        except:
            pass

@dp.message(F.text == "/users")
async def list_users(msg: Message):
    if msg.from_user.id not in ADMIN_USERS:
        return

    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT user_id, username, first_name, last_name, joined_at
            FROM users
            ORDER BY joined_at DESC
        """)

    if not rows:
        await msg.answer("Користувачів ще немає.")
        return

    text = "📋 Користувачі бота:\n\n"
    for r in rows:
        username = f"@{r['username']}" if r['username'] else "(без username)"
        name = f"{r['first_name']} {r['last_name'] or ''}".strip()
        joined = r['joined_at'].astimezone(TIMEZONE).strftime("%d.%m.%Y %H:%M")
        text += f"• <b>{name}</b> {username}\n  ID: <code>{r['user_id']}</code>\n  Приєднався: {joined}\n\n"

    await msg.answer(text, parse_mode="HTML")


@dp.callback_query(F.data.in_({"last_7", "last_30", "last_90"}))
async def last_fixed_days(cb: CallbackQuery):
    days_map = {"last_7": 7, "last_30": 30, "last_90": 90}
    days = days_map[cb.data]
    now = datetime.now(TIMEZONE)
    start = now - timedelta(days=days)
    total, count, avg = await calculate_period(cb.from_user.id, start)
    await safe_edit_message(cb,
        f"За останні {days} днів\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн",
        reply_markup=stats_keyboard()
    )


# === Кастомні періоди ===
@dp.callback_query(F.data == "custom_range")
async def custom_range_start(cb: CallbackQuery, state: FSMContext):
    await state.set_state(CustomStats.waiting_range)
    await cb.message.answer(
        "Введи діапазон у форматі:\n"
        "<code>ДД.ММ.РРРР - ДД.ММ.РРРР</code>\n"
        "Наприклад: <code>01.10.2025 - 30.12.2025</code>",
        parse_mode="HTML"
    )


@dp.message(CustomStats.waiting_range)
async def process_custom_range(msg: Message, state: FSMContext):
    pattern = r"^(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})$"
    match = re.fullmatch(pattern, msg.text.strip())
    if not match:
        await msg.answer("Неправильний формат. Використовуй: ДД.ММ.РРРР - ДД.ММ.РРРР")
        return
    try:
        d1_str, d2_str = match.groups()
        day1, mon1, year1 = map(int, d1_str.split("."))
        day2, mon2, year2 = map(int, d2_str.split("."))
        start = datetime(year1, mon1, day1)
        end = datetime(year2, mon2, day2, 23, 59, 59) + timedelta(seconds=1)
        if start >= end:
            await msg.answer("Початкова дата має бути раніше кінцевої")
            return
    except ValueError:
        await msg.answer("Некоректна дата")
        return

    total, count, avg = await calculate_period(msg.from_user.id, start, end)
    await msg.answer(
        f"Статистика з {d1_str} по {d2_str}\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн"
    )
    await state.clear()


@dp.callback_query(F.data == "last_n_days")
async def last_n_days_start(cb: CallbackQuery, state: FSMContext):
    await state.set_state(CustomStats.waiting_days)
    await cb.message.answer("Введи кількість днів (наприклад: 45):")


@dp.message(CustomStats.waiting_days)
async def process_last_n_days(msg: Message, state: FSMContext):
    if not re.fullmatch(r"\d+", msg.text.strip()):
        await msg.answer("Введи тільки число")
        return
    days = int(msg.text.strip())
    if days <= 0 or days > MAX_CUSTOM_DAYS:
        await msg.answer(f"Кількість днів має бути від 1 до {MAX_CUSTOM_DAYS}")
        return
    now = datetime.now(TIMEZONE)
    start = now - timedelta(days=days)
    total, count, avg = await calculate_period(msg.from_user.id, start)
    await msg.answer(
        f"За останні {days} днів\n\n"
        f"Сума: {total} грн\n"
        f"Днів: {count}\n"
        f"Середній чек: {avg:.2f} грн"
    )
    await state.clear()


# === Редагування ===
@dp.callback_query(F.data == "edit")
async def edit_start(cb: CallbackQuery, state: FSMContext):
    await ensure_user_table(cb.from_user.id)
    await cb.message.edit_text(
        "✏️ Обери, за яку дату додати/змінити суму чайових:",
        reply_markup=date_choice_keyboard()
    )
    await state.set_state(EditSum.waiting_for_date)  # чекаємо або кнопку, або текст


# Обробка кнопок "Сьогодні" і "Вчора"
@dp.callback_query(F.data.in_({"date_today", "date_yesterday"}))
async def process_quick_date(cb: CallbackQuery, state: FSMContext):
    from datetime import datetime, timedelta
    now = datetime.now(TIMEZONE)
    if cb.data == "date_today":
        selected_date = now.date()
        date_str = now.strftime("%d.%m.%Y")
    else:  # date_yesterday
        selected_date = (now - timedelta(days=1)).date()
        date_str = (now - timedelta(days=1)).strftime("%d.%m.%Y")

    selected_dt = datetime.combine(selected_date, datetime.min.time())

    await state.update_data(selected_date=selected_dt)
    await state.set_state(EditSum.waiting_for_sum)

    await cb.message.edit_text(
        f"📅 Обрано дату: <b>{date_str}</b>\n\nВведи суму чайових (лише число, може бути від'ємним):",
        parse_mode="HTML"
    )


# Кнопка "Ввести вручну" — просто підказка
@dp.callback_query(F.data == "date_manual")
async def date_manual_prompt(cb: CallbackQuery):
    await cb.message.edit_text(
        "✍️ Введи дату у форматі <code>ДД.ММ.РРРР</code>\n"
        "Наприклад: <code>04.01.2026</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Назад до вибору", callback_data="edit")]
        ])
    )


@dp.message(EditSum.waiting_for_date)
async def process_date(msg: Message, state: FSMContext):
    text = msg.text.strip()
    if not re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        await msg.answer("❌ Неправильний формат. Використовуй ДД.ММ.РРРР")
        return

    try:
        day, month, year = map(int, text.split("."))
        selected_date = datetime(year, month, day)
    except ValueError:
        await msg.answer("❌ Некоректна дата")
        return

    await state.update_data(selected_date=selected_date)
    await state.set_state(EditSum.waiting_for_sum)
    await msg.answer(f"📅 Дата {text}\n\nВведи суму чайових (лише число):")


@dp.message(EditSum.waiting_for_sum)
async def process_sum(msg: Message, state: FSMContext):
    if not re.fullmatch(r"-?\d+", msg.text.strip()):
        await msg.answer("Введи тільки число")
        return
    tips = int(msg.text.strip())
    data = await state.get_data()
    selected_date = data["selected_date"]
    user_id = msg.from_user.id

    old_record = await upsert_tips(user_id, selected_date, tips)

    if old_record:
        await msg.answer(
            f"Сума на {selected_date.strftime('%d.%m.%Y')} змінена:\n"
            f"{old_record['tips']} → {tips} грн"
        )
    else:
        await msg.answer(f"Додано {tips} грн на {selected_date.strftime('%d.%m.%Y')}")

    await state.clear()


# ================== RUN ==================
async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())